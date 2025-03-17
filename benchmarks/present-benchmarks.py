# present captures benchmarks

import json
import os
import pandas as pd
from glob import glob
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import argparse

def analyze_and_present_results(results_dir):
    # Get all JSON files in the results directory
    json_files = glob(os.path.join(results_dir, "*.json"))
    
    if not json_files:
        print(f"No JSON files found in directory: {results_dir}")
        return
    
    # List to store all results
    results_data = []
    
    # Read each JSON file
    for json_file in json_files:
        with open(json_file, 'r') as f:
            data = json.load(f)
            
            # Calculate output throughput (tokens per second)
            output_throughput = (data['total_output_tokens'] * 1000) / data['mean_tpot_ms']
            
            # Calculate token ratio (input/output)
            token_ratio = round(data['total_input_tokens'] / data['total_output_tokens'], 2)
            
            # Extract the metrics we want and round numeric values
            result = {
                'total_input_tokens': data['total_input_tokens'],
                'total_output_tokens': data['total_output_tokens'],
                'token_ratio': token_ratio,
                'concurrency': int(data['max_concurrency']),  # Ensure concurrency is integer
                'mean_ttft_ms': round(data['mean_ttft_ms'], 2),
                'mean_tpot_ms': round(data['mean_tpot_ms'], 2),
                'mean_itl_ms': round(data['mean_itl_ms'], 2),
                'output_throughput': round(data['output_throughput'], 2)
            }
            results_data.append(result)
    
    # Convert to DataFrame
    df = pd.DataFrame(results_data)
    
    # Sort DataFrame by concurrency only (as integer)
    df = df.sort_values('concurrency')
    
    # Create pivot tables for each metric
    metrics = ['mean_ttft_ms', 'mean_tpot_ms', 'mean_itl_ms', 'output_throughput']
    pivot_tables = {}
    
    for metric in metrics:
        pivot = pd.pivot_table(
            df,
            values=metric,
            index=['total_input_tokens', 'total_output_tokens'],
            columns='concurrency',
            aggfunc='mean'
        )
        # Round values in pivot table
        pivot = pivot.round(2)
        pivot_tables[metric] = pivot
    
    # Create output filename using directory name
    dir_name = os.path.basename(os.path.normpath(results_dir))
    output_file = f"{dir_name}_benchmark_results.xlsx"
    
    # Create Excel writer object
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write raw data first
        df.to_excel(writer, sheet_name='Raw Data', index=False, float_format="%.2f")
        
        # Write pivot tables
        for metric, pivot in pivot_tables.items():
            sheet_name = metric.replace('mean_', '').replace('_ms', '')
            pivot.to_excel(writer, sheet_name=sheet_name, float_format="%.2f")
            
            # Auto-adjust columns width
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(pivot.columns):
                # Get the maximum length in the column
                max_length = max(
                    pivot[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                # Get the column letter
                column_letter = get_column_letter(idx + 2)  # +2 because of index column
                # Set the column width
                worksheet.column_dimensions[column_letter].width = max_length + 2
            
            # Adjust the first column (index)
            worksheet.column_dimensions['A'].width = max(
                len(str(idx)) for idx in pivot.index
            ) + 2
        
        # Auto-adjust Raw Data sheet
        worksheet = writer.sheets['Raw Data']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            column_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = max_length + 2
    
    print(f"Results have been written to {output_file}")

    # Create summary statistics
    print("\nSummary Statistics:")
    print("-" * 50)
    for input_tokens in df['total_input_tokens'].unique():
        for output_tokens in df[df['total_input_tokens'] == input_tokens]['total_output_tokens'].unique():
            subset = df[
                (df['total_input_tokens'] == input_tokens) & 
                (df['total_output_tokens'] == output_tokens)
            ]
            print(f"\nInput Tokens: {input_tokens}, Output Tokens: {output_tokens}")
            print(f"Average TTFT: {subset['mean_ttft_ms'].mean():.2f} ms")
            print(f"Average TPOT: {subset['mean_tpot_ms'].mean():.2f} ms")
            print(f"Average ITL: {subset['mean_itl_ms'].mean():.2f} ms")
            print(f"Average Output Throughput: {subset['output_throughput'].mean():.2f} tokens/sec")
            print(f"Min/Max Concurrency: {subset['concurrency'].min()}/{subset['concurrency'].max()}")

def main():
    parser = argparse.ArgumentParser(description='Analyze benchmark results from a specified directory')
    parser.add_argument('results_dir', type=str, help='Directory containing the JSON result files')
    args = parser.parse_args()
    
    analyze_and_present_results(args.results_dir)

if __name__ == "__main__":
    main()